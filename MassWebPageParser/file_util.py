import os  # listing directories etc.
import openpyxl  # to save to a file for further upload
import codecs  # to read html
from pathlib import Path  # to create files


class FileUtil:

    descs_local_dir = ''
    pics_local_dir = ''
    excel_local_dir = ''
    excel_local_link = ''

    worksheet_name = ''
    img_name_search_tag = ''

    # ADDITIONAL METHODS

    @staticmethod
    def add_tailing_slash(local_url):
        if local_url[-1] not in ['/', '\\']:
            print()
            print("No tailing slash found, adding...")
            local_url = os.path.join(local_url, '')
        else:
            pass
        return local_url

    @staticmethod
    def check_file_creation(name, file_dir):
        if os.path.isfile(file_dir):  # check if the new html file is created
            print()
            print('The file named ' + name + ' is successfully created.')
        else:
            print()
            print('Some problems with empty html file creation occurred, please check.')

    # MAIN METHODS

    @staticmethod
    def insert_dirs():
        # saves the inserted dir values to class variables

        print("Please state the directory to store description files in:")
        FileUtil.descs_local_dir = FileUtil.add_tailing_slash(input())

        print("Now state the directory to store downloaded images in:")
        FileUtil.pics_local_dir = FileUtil.add_tailing_slash(input())

    @staticmethod
    def insert_search_params():
        print()
        print('Now state the tag in which to search for the image name which is the product name (i.e. "h1"):')
        FileUtil.img_name_search_tag = input()

        print("Please state the name of the excel worksheet in use:")
        FileUtil.worksheet_name = input()

        print()

    @staticmethod
    def get_excel_file():

        print("Do you have an excel file for upload ready (y/n)?")
        answer = input()

        while answer not in ['y', 'Y', 'n', 'N']:
            print("Not a valid answer, try again.")
            print("Do you have an excel file for upload ready (y/n)?")
            answer = input()

        print()

        if answer in ['y', 'Y']:
            print(
                "Please paste the complete absolute link to an excel file:")
            print(
                "In this file descriptions and relative links to images will be stored to be uploaded to your e-store.")
            FileUtil.excel_local_link = input()

        elif answer in ['n', 'N']:
            print("Please give the new excel file a name:")
            excel_file_name = input()

            print("Please state the directory in which you want the excel file to be stored:")
            FileUtil.excel_local_dir = FileUtil.add_tailing_slash(input())

            FileUtil.excel_local_link = FileUtil.excel_local_dir + excel_file_name + '.xls'
            Path(FileUtil.excel_local_link).touch()  # creates an excel file

            # call the checker function to find if the new file was created
            FileUtil.check_file_creation(excel_file_name, FileUtil.excel_local_link)

            print("Now rename your excel file and fill it with relevant data.")

    # SENDING METHODS

    @staticmethod
    def send_descs_to_excel():
        desc_files = os.listdir(FileUtil.descs_local_dir)

        wb = openpyxl.load_workbook(FileUtil.excel_local_link)  # opening xls file as an object
        ws = wb[FileUtil.worksheet_name]

        for desc_file in desc_files:
            print('Next file: ' + desc_file)  # looping through all files in the directory

            # open html to check

            f = codecs.open(FileUtil.descs_local_dir + desc_file, 'r', 'utf-8')
            f_contents = f.read()

            if len(str(f_contents)) > 0:
                print('Found some text in file!')
            else:
                print('Sorry, no text found.')

            # searching for file name occurrence in cells
            # the code below is a sample, do NOT paste it in your project as is
            """
            for row in ws.iter_rows(min_row=2, max_col=1, max_row=273):
                for cell in row:
                    if str(desc_file[0:-5]).upper() in str(cell.value).upper():
                        print("Match for SKU " + desc_file + " found" + " on row " + str(cell.row))

                        # writing to each cell that has an occurrence

                        ws.cell(column=3, row=cell.row).value = f_contents  # filling the C column with descs
                        print('Row ' + str(cell.row) + ' in column C is filled with description text.')

                        f.close()
                        print()
                    else:
                        pass
            """

        print()
        wb.save(FileUtil.excel_local_link)  # saving the values
        wb.close()

    @staticmethod
    def send_pics_links_to_excel():
        wb = openpyxl.load_workbook(FileUtil.excel_local_link)  # opening xls file as an object
        ws = wb[FileUtil.worksheet_name]

        inner_dict = {}

        for pic_file in os.listdir(FileUtil.pics_local_dir):

            # searching for pic name occurrence in cells
            # the code below is a sample, do NOT paste it in your project as is

            """
            for row in ws.iter_rows(min_row=2, max_col=1, max_row=273):  # each row in column
                for cell in row:  # each cell in row
                    if str(pic_file[0:-4]).upper() in str(cell.value).upper():
                        inner_dict[str(cell.value)] = pic_file
                    else:
                        pass
            """

        # the code below is a sample, do NOT paste it in your project as is
        """
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=273):
            for cell in row:
                for key, value in inner_dict.items():
                    if key == cell.value:
                        print("Found match, writing results to row " + str(cell.row))
                        ws.cell(column=4, row=cell.row).value = (
                                '/upload/pics/blabla/' + value)  # filling the column with pic links
                        ws.cell(column=5, row=cell.row).value = (
                                '/upload/pics/blabla/' + value)
                    else:
                        pass
        """

        wb.save(FileUtil.excel_local_link)  # saving the values
